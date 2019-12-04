#!/usr/bin/python3
from openpyxl import *
res = dict()
data = []
total = []
def fill_dic(grade, idx):
    if total[idx] < 40:
        preidx = idx - 1
        while preidx >= 0:
            if total[preidx] >= 40:
                idx = preidx
                break
            else :
                preidx -= 1
        if preidx == -1:
            res[grade] = -1
            return

    if total[idx] != total[idx + 1]:
        res[grade] = idx
    else :
        preidx = idx - 1
        while preidx >= 0:
            if total[preidx] == total[idx]:
                idx = preidx
                preidx -= 1
            else:
                res[grade] = preidx
                break
        if res.get(grade) is None:
            res[grade] = -1


def calc_ratio():
    wb = load_workbook(filename='student.xlsx')
    sheet = wb['Sheet1']

    i=2
    while sheet.cell(row=i, column=1).value is not None:
        midterm = sheet.cell(row = i, column = 3).value
        final = sheet.cell(row = i, column = 4).value
        hw = sheet.cell(row = i, column = 5).value
        attendance = sheet.cell(row = i, column = 6).value
        cell_total = midterm * 0.3 + final * 0.35 + hw * 0.34 + attendance
        sheet.cell(row = i, column = 7, value="{}".format(cell_total))
        total.append(cell_total)
        i += 1
    wb.save('student.xlsx')
    total.sort(reverse=True)

    data = [['C+', 0.85], ['B', 0.7], ['B+', 0.5], ['A', 0.3], ['A+', 0.15]]
    for item in data:
        fill_dic(item[0], int(len(total) * item[1]) - 1)


def fill():
    wb = load_workbook(filename='student.xlsx')
    sheet = wb['Sheet1']
    j=2
    while sheet.cell(row=j, column=1).value is not None:
        t = float(sheet.cell(row=j, column=7).value)
        grade=''
        if res['A+'] != -1 and t >= total[res['A+']]:
            grade = 'A+'
        elif res['A'] != -1 and t >= total[res['A']]:
            grade = 'A'
        elif res['B+'] != -1 and t >= total[res['B+']]:
            grade = 'B+'
        elif res['B'] != -1 and t >= total[res['B']]:
            grade = 'B'
        elif res['C+'] != -1 and t >= total[res['C+']]:
            grade = 'C+'
        elif t >= 40:
            grade = 'C'
        else :
            grade = 'F'
        sheet.cell(row=j, column=8, value="{}".format(grade))
        j += 1
    wb.save('student.xlsx')


if __name__ == '__main__':
    calc_ratio()
    fill()
